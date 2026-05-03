#!/usr/bin/env python3
"""Merge KESSES STOCK .xlsx pharma sheets into drugs.csv for Odoo initializer.

Sets ``product.product`` fields from ``ampath_billing`` (see addon models):

  - ``x_concept_source`` — fixed ``local`` (only Selection option defined today).
  - ``x_concept_code`` — from Excel ``Concept ID``, else ``Local Mapping``, else ``Drug ID``.
  - ``x_drug_strength`` — regex extract from product ``name`` (e.g. 500MG); Excel row optional.

Excel tabs used: Pharmaceuticals Stock, Oncology, KEMSA stocks received_latest.

Run from repo root::

  python3 scripts/enrich_drugs_kesses.py

Requires: pip install openpyxl
"""
from __future__ import annotations

import csv
import math
import re
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / 'KESSES STOCK .xlsx'
DRUGS_CSV = ROOT / 'distro/config/odoo/initializer_config/product_variant/drugs.csv'

STRENGTH_RE = re.compile(
    r'\b(\d+(?:\.\d+)?)\s*(MG(?:\/ML)?|MCG|µG|UG|ML|GM|G|%|IU)\b',
    re.I,
)


def norm(s):
    if s is None:
        return ''
    s = str(s).strip().upper()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^A-Z0-9\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def tokens(s):
    return set(norm(s).split()) - {'AND', 'THE', 'FOR', 'WITH', 'TAB', 'TABS', 'TABLET', 'TABLETS', 'CAP', 'CAPS', 'MG', 'ML', 'GM'}


def strength_from_text(name):
    if not name:
        return ''
    m = STRENGTH_RE.search(str(name))
    if m:
        return (m.group(1) + m.group(2)).upper().replace('µ', 'U')
    return ''


def concept_parts(concept_id, drug_id, local_map):
    for v in (concept_id, local_map, drug_id):
        if v is None or v == '':
            continue
        if isinstance(v, float) and math.isnan(v):
            continue
        s = str(v).strip()
        if s.endswith('.0') and s[:-2].replace('.', '').isdigit():
            s = s[:-2]
        if s and s.lower() != 'none':
            return s
    return ''


def sheet_records(ws):
    rows = list(ws.iter_rows(values_only=True))
    hdr_ix = None
    headers = None
    for i, row in enumerate(rows):
        if not row:
            continue
        row_lower = [str(c).lower() if c is not None else '' for c in row]
        line = ' '.join(row_lower)
        if 'product' in line and ('concept' in line or 'description' in line):
            hdr_ix = i
            headers = [str(c).strip() if c is not None else '' for c in row]
            break
    if hdr_ix is None or not headers:
        return []

    def find_col(candidates):
        for j, h in enumerate(headers):
            hclean = h.strip()
            for c in candidates:
                if hclean.lower() == c.lower():
                    return j
        return None

    pi = find_col(('Product', 'Description'))
    ci = find_col(('Concept ID', 'CONCEPT ID'))
    di = find_col(('Drug ID', 'DRUG ID'))
    lmi = find_col(('Local Mapping',))
    if pi is None:
        return []

    recs = []
    for row in rows[hdr_ix + 1 :]:
        if not row or pi >= len(row):
            continue
        prod = row[pi]
        if prod is None:
            continue
        prod = str(prod).strip()
        if not prod:
            continue

        def gv(ix):
            if ix is None or ix >= len(row):
                return None
            return row[ix]

        code = concept_parts(gv(ci), gv(di), gv(lmi))
        recs.append({'product': prod, 'code': code, 'norm': norm(prod), 'tokens': tokens(prod)})
    return recs


def best_match(name, records):
    n = norm(name)
    if not n:
        return None
    nt = tokens(name)
    best = None
    best_score = 0.0
    for r in records:
        if r['norm'] == n:
            return r
        if r['norm'] and (r['norm'] in n or n in r['norm']):
            score = min(len(r['norm']), len(n)) / max(len(r['norm']), len(n), 1)
            if score > best_score:
                best_score = score
                best = r
        if nt and r['tokens']:
            inter = len(nt & r['tokens'])
            union = len(nt | r['tokens']) or 1
            j = inter / union
            if j > best_score and inter >= 2:
                best_score = j
                best = r
    if best_score >= 0.35:
        return best
    return None


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print('Install: pip install openpyxl', file=sys.stderr)
        return 1

    if not XLSX.is_file():
        print('Missing', XLSX, file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    records: list[dict] = []
    for sn in ('Pharmaceuticals Stock', 'Oncology', 'KEMSA stocks received_latest'):
        if sn in wb.sheetnames:
            records.extend(sheet_records(wb[sn]))
    wb.close()

    by_norm: dict[str, dict] = {}
    for r in records:
        k = r['norm']
        if k not in by_norm or (r['code'] and not by_norm[k].get('code')):
            by_norm[k] = r
    records = list(by_norm.values())

    with DRUGS_CSV.open(newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames or [])
        rows = list(reader)

    for col in ('x_concept_source', 'x_concept_code', 'x_drug_strength'):
        if col not in fieldnames:
            fieldnames.append(col)

    for row in rows:
        name = (row.get('name') or '').strip()
        row['x_concept_source'] = 'local'
        row['x_drug_strength'] = strength_from_text(name)
        m = best_match(name, records)
        row['x_concept_code'] = (m.get('code') or '') if m else ''

    with DRUGS_CSV.open('w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        w.writeheader()
        w.writerows(rows)

    n_code = sum(1 for r in rows if (r.get('x_concept_code') or '').strip())
    n_str = sum(1 for r in rows if (r.get('x_drug_strength') or '').strip())
    print(f'Wrote {DRUGS_CSV}: {len(rows)} rows, x_concept_code set={n_code}, x_drug_strength set={n_str}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
