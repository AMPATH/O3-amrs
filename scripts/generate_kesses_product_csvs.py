#!/usr/bin/env python3
"""Create Odoo initializer CSVs from ``KESSES STOCK .xlsx`` (one file per sheet).

Outputs **only products that are not already present** in
``distro/config/odoo/initializer_config/product_variant/*.csv`` (matched by normalized name).

* **Drug** sheets (pharma, oncology, KEMSA): same columns as ``drugs.csv`` — stockable product,
  drug category, concept/strength fields when available.
* **Supply / equipment** sheets: stockable **Materials** category, generic UOM (same id as drugs csv).

Run from repo root::

    pip install openpyxl
    python3 scripts/generate_kesses_product_csvs.py

"""
from __future__ import annotations

import csv
import math
import re
import sys
import uuid
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / 'scripts'))
from pharmaceutical_uom_rules import infer_pharmaceutical_uom  # noqa: E402
from product_name_canonical import canonical_drug_key  # noqa: E402
VARIANT_DIR = ROOT / 'distro/config/odoo/initializer_config/product_variant'
XLSX = ROOT / 'KESSES STOCK .xlsx'

# product.category.xml ids
CATEGORY_DRUG = 'init.categ_products_drug_orders'
CATEGORY_MATERIALS = 'init.categ_products_materials_orders'

# Materials / non-drug sheets: generic count UOM from initializer bundle
DEFAULT_UOM = 'init.a8a0630a-1350-11df-a1f1-0026b9348838'
DEFAULT_UOM_PO = 'init.a8a0630a-1350-11df-a1f1-0026b9348838'

NS_KESSES = uuid.UUID('b7e75440-2e9d-5d6e-9e1f-a4c6c6e8f0d1')

STRENGTH_RE = re.compile(
    r'\b(\d+(?:\.\d+)?)\s*(MG(?:\/ML)?|MCG|µG|UG|ML|GM|G|%|IU)\b',
    re.I,
)


def strength_from_text(name: str) -> str:
    if not name:
        return ''
    m = STRENGTH_RE.search(str(name))
    if m:
        return (m.group(1) + m.group(2)).upper().replace('µ', 'U')
    return ''


def concept_parts(concept_id, drug_id, local_map):
    for v in (concept_id, local_map, drug_id):
        if v is None or v == '':
            continue
        if isinstance(v, float) and math.isnan(v):
            continue
        s = str(v).strip()
        if s.endswith('.0') and s[:-2].replace('.', '').isdigit():
            s = s[:-2]
        if s and s.lower() != 'none':
            return s
    return ''


def load_existing_product_names() -> set[str]:
    """All product names from existing CSVs, using ``canonical_drug_key`` for matching.

    Same key as ``dedupe_kesses_drugs.py`` so TAB/TABLET/CAPSULE variants do not
    reappear under KESSES when ``drugs.csv`` already lists the item.
    """
    out: set[str] = set()
    for p in sorted(VARIANT_DIR.glob('*.csv')):
        try:
            with p.open(newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    n = (row.get('name') or '').strip().strip('"')
                    if n:
                        out.add(canonical_drug_key(n))
        except OSError:
            pass
    return out


def stable_xml_id(sheet_slug: str, product_code: str, name: str) -> str:
    key = f'kesses|{sheet_slug}|{product_code}|{name}'
    return f'init.{uuid.uuid5(NS_KESSES, key)}'


def float_or_none(v):
    if v is None or v == '':
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def row_dict(headers: list[str], row: tuple) -> dict:
    d = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        key = str(h).strip()
        val = row[i] if i < len(row) else None
        d[key] = val
    return d


def find_header_row(rows: list[tuple]) -> tuple[int, list[str]] | tuple[None, None]:
    for i, row in enumerate(rows[:25]):
        if not row:
            continue
        labels = [str(c).strip() if c is not None else '' for c in row]
        joined = ' '.join(x.lower() for x in labels if x)
        if 'product' in joined or 'description' in joined:
            if any('price' in x.lower() for x in labels if x) or 'kemsa' in joined:
                return i, labels
            if any(x.lower() in ('product', 'description') for x in labels if x):
                return i, labels
    return None, None


def pick_price_buy(rd: dict) -> tuple[float | None, float | None]:
    """(lst_price, standard_price) from messy excel headers."""
    sell_keys = (
        'Selling Price per item',
        'SELLING PRICE',
        'Selling Price',
        'Price per Item',
        'Price',
        'UNIT PRICE',
    )
    buy_keys = (
        'Buying Price per item',
        'PRICE PER ITEM (KSHS)',
        'Buying Price',
    )
    lst = None
    std = None
    for k, v in rd.items():
        kl = k.strip()
        if kl in sell_keys or kl.upper() in [s.upper() for s in sell_keys]:
            lst = float_or_none(v)
            if lst is not None:
                break
    for k, v in rd.items():
        kl = k.strip()
        if kl in buy_keys or 'BUY' in kl.upper():
            std = float_or_none(v)
            if std is not None:
                break
    if lst is None:
        for k, v in rd.items():
            if 'price' in k.lower() and 'buy' not in k.lower():
                x = float_or_none(v)
                if x is not None:
                    lst = x
                    break
    if std is None and lst is not None:
        std = round(lst * 0.8, 2)
    if lst is None:
        lst = 0.0
    if std is None:
        std = 0.0
    return lst, std


def sheet_slug(name: str) -> str:
    s = re.sub(r'[^a-zA-Z0-9]+', '_', name.strip())
    return s.strip('_').lower() or 'sheet'


def parse_sheet(ws, sheet_name: str, is_drug_sheet: bool):
    rows = list(ws.iter_rows(values_only=True))
    hdr_ix, headers = find_header_row(rows)
    if hdr_ix is None or not headers:
        return []
    headers = [str(h).strip() if h is not None else '' for h in headers]
    out = []
    for row in rows[hdr_ix + 1 :]:
        if not row:
            continue
        rd = row_dict(headers, row)
        prod = rd.get('Product') or rd.get('Description') or rd.get('PRODUCT')
        if prod is None:
            continue
        prod = str(prod).strip()
        if not prod or prod.lower() == 'description':
            continue
        code = rd.get('Product_code') or rd.get('Product code') or ''
        if code is not None:
            code = str(code).strip()
        lst, std = pick_price_buy(rd)
        concept_code = ''
        if is_drug_sheet:
            concept_code = concept_parts(
                rd.get('Concept ID'),
                rd.get('Drug ID'),
                rd.get('Local Mapping'),
            )
        out.append(
            {
                'product_code': code or '-',
                'name': prod,
                'lst_price': lst,
                'standard_price': std,
                'x_concept_code': concept_code,
                'x_drug_strength': strength_from_text(prod),
            }
        )
    return out


DRUG_SHEETS = {
    'Pharmaceuticals Stock',
    'Oncology',
    'KEMSA stocks received_latest',
}


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print('pip install openpyxl', file=sys.stderr)
        return 1

    if not XLSX.is_file():
        print('Missing', XLSX, file=sys.stderr)
        return 1

    existing = load_existing_product_names()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    total_new = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_drug = sheet_name in DRUG_SHEETS
        records = parse_sheet(ws, sheet_name, is_drug)
        slug = sheet_slug(sheet_name)
        new_rows: list[dict] = []
        for rec in records:
            ck = canonical_drug_key(rec['name'])
            if not ck or ck in existing:
                continue
            existing.add(ck)
            pid = stable_xml_id(slug, rec['product_code'], rec['name'])
            lst = rec['lst_price'] if rec['lst_price'] is not None else 0.0
            std = rec['standard_price'] if rec['standard_price'] is not None else 0.0
            drug_uom = infer_pharmaceutical_uom(rec['name'])
            if is_drug:
                new_rows.append(
                    {
                        'id': pid,
                        'name': rec['name'],
                        'categ_id/id': CATEGORY_DRUG,
                        'type': 'product',
                        'tracking': 'lot',
                        'use_expiration_date': 'True',
                        'uom_id/id': drug_uom,
                        'uom_po_id/id': drug_uom,
                        'invoice_policy': 'Ordered quantities',
                        'lst_price': f'{lst:.2f}',
                        'standard_price': f'{std:.2f}',
                        'x_concept_source': 'local',
                        'x_concept_code': rec['x_concept_code'] or '',
                        'x_drug_strength': rec['x_drug_strength'] or '',
                    }
                )
            else:
                new_rows.append(
                    {
                        'id': pid,
                        'name': rec['name'],
                        'categ_id/id': CATEGORY_MATERIALS,
                        'type': 'product',
                        'tracking': 'lot',
                        'use_expiration_date': 'True',
                        'uom_id/id': DEFAULT_UOM,
                        'uom_po_id/id': DEFAULT_UOM_PO,
                        'invoice_policy': 'Ordered quantities',
                        'lst_price': f'{lst:.2f}',
                        'standard_price': f'{std:.2f}',
                    }
                )
        ws_title = f'kesses_{slug}.csv'
        out_path = VARIANT_DIR / ws_title
        if new_rows:
            if is_drug:
                fieldnames = [
                    'id',
                    'name',
                    'categ_id/id',
                    'type',
                    'tracking',
                    'use_expiration_date',
                    'uom_id/id',
                    'uom_po_id/id',
                    'invoice_policy',
                    'lst_price',
                    'standard_price',
                    'x_concept_source',
                    'x_concept_code',
                    'x_drug_strength',
                ]
            else:
                fieldnames = [
                    'id',
                    'name',
                    'categ_id/id',
                    'type',
                    'tracking',
                    'use_expiration_date',
                    'uom_id/id',
                    'uom_po_id/id',
                    'invoice_policy',
                    'lst_price',
                    'standard_price',
                ]
            with out_path.open('w', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
                w.writeheader()
                w.writerows(new_rows)
            total_new += len(new_rows)
            print(f'{sheet_name}: +{len(new_rows)} -> {out_path.name}')
        else:
            print(f'{sheet_name}: 0 new (all rows already in initializer CSVs)')

    wb.close()
    print(f'Total new products: {total_new}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
