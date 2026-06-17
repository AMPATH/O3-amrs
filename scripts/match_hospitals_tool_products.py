#!/usr/bin/env python3
"""Match KEMSA hospitals ordering tool products to Odoo initializer CSVs.

Parses the KEMSA LMIS hospitals Excel (``Ordering Tool`` sheet), matches each
product by canonical name against ``product_variant/*.csv``, and writes analysis
reports (matched, missing, ambiguous, reverse gap).

Run from repo root::

    pip install openpyxl
    python3 scripts/match_hospitals_tool_products.py \\
      --xlsx "/path/to/hospitals-tool-latest.xlsx" \\
      --out reports/hospitals-tool-match

"""
from __future__ import annotations

import argparse
import csv
import math
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / 'scripts'))

from product_name_canonical import canonical_drug_key  # noqa: E402

DEFAULT_XLSX = Path(
    '/Users/emmanuelnyachoke/Downloads/hospitals-tool-latest (7).xlsx'
)
DEFAULT_VARIANT_DIR = ROOT / 'distro/config/odoo/initializer_config/product_variant'
DEFAULT_OUT = ROOT / 'reports/hospitals-tool-match'

TRACER_RE = re.compile(r'\s*\*\*\*\s*Tracer\s*\*\*\*\s*', re.I)

SECTION_TO_CSV: dict[str, str] = {
    'PHARMACEUTICALS': 'kesses_pharmaceuticals_stock.csv',
    'SSD PHARMACEUTICALS': 'kesses_pharmaceuticals_stock.csv',
    'ONCOLOGY': 'kesses_oncology.csv',
    'RENAL': 'kesses_renal.csv',
    'NON-PHARMACEUTICALS': 'kesses_non_pharmaceuticals.csv',
    'LAB PRODUCTS': 'kesses_lab_products.csv',
    'LINEN PRODUCTS': 'kesses_linen_products.csv',
    'X-RAY': 'kesses_xray.csv',
    'DENTAL PRODUCTS': 'kesses_dental_products.csv',
    'PUBLIC HEALTH PRODUCTS': 'kesses_public_health.csv',
    'BASIC EQUIPMENT': 'kesses_basic_equipment.csv',
    'EQUIPMENT': 'kesses_equipments.csv',
    'CHV KITS': 'kesses_chv.csv',
    'ORTHOPAEDIC EQUIPMENT': 'kesses_orthopaedic.csv',
    'OTHER': 'kesses_non_pharmaceuticals.csv',
}

STOP_TOKENS = frozenset(
    {
        'AND',
        'THE',
        'FOR',
        'WITH',
        'TAB',
        'TABS',
        'TABLET',
        'TABLETS',
        'CAP',
        'CAPS',
        'MG',
        'ML',
        'GM',
    }
)


def norm_tokens(s: str) -> set[str]:
    if not s:
        return set()
    t = str(s).strip().upper()
    t = unicodedata.normalize('NFKD', t)
    t = ''.join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r'[^A-Z0-9\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    return set(t.split()) - STOP_TOKENS


def strip_tracer(name: str) -> tuple[str, bool]:
    if not name:
        return '', False
    cleaned = TRACER_RE.sub(' ', str(name))
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned, cleaned != str(name).strip()


def float_or_str(v) -> str:
    if v is None or v == '':
        return ''
    if isinstance(v, float) and math.isnan(v):
        return ''
    if isinstance(v, (int, float)):
        if float(v) == int(v):
            return str(int(v))
        return str(v)
    return str(v).strip()


def parse_excel(xlsx: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit('Requires openpyxl: pip install openpyxl') from exc

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if 'Ordering Tool' not in wb.sheetnames:
        raise SystemExit(f'Sheet "Ordering Tool" not found in {xlsx}')
    ws = wb['Ordering Tool']

    products: list[dict] = []
    current_section = 'UNKNOWN'

    for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_num <= 5:
            continue
        code = row[2] if len(row) > 2 else None
        name = row[3] if len(row) > 3 else None
        category = row[4] if len(row) > 4 else None
        pack_size = row[5] if len(row) > 5 else None
        price = row[6] if len(row) > 6 else None

        if not code or not str(code).strip():
            continue
        code_s = str(code).strip()

        if name is None and category is None:
            if not code_s.upper().startswith('TOTAL'):
                current_section = code_s.upper()
            continue
        if name is None:
            continue

        name_s = str(name).strip()
        if not name_s:
            continue

        match_name, is_tracer = strip_tracer(name_s)
        products.append(
            {
                'kemsa_code': code_s,
                'name': name_s,
                'match_name': match_name,
                'is_tracer': is_tracer,
                'category': str(category).strip() if category else '',
                'pack_size': str(pack_size).strip() if pack_size else '',
                'price': float_or_str(price),
                'section': current_section,
                'excel_row': row_num,
                'key': canonical_drug_key(match_name),
            }
        )

    wb.close()
    return products


def load_csv_index(variant_dir: Path) -> tuple[dict[str, list[dict]], list[dict]]:
    by_key: dict[str, list[dict]] = {}
    all_rows: list[dict] = []

    for path in sorted(variant_dir.glob('*.csv')):
        with path.open(newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = (row.get('name') or '').strip().strip('"')
                if not name:
                    continue
                rec = {
                    'id': (row.get('id') or '').strip().strip('"'),
                    'name': name,
                    'file': path.name,
                    'lst_price': (row.get('lst_price') or '').strip().strip('"'),
                    'default_code': (row.get('default_code') or '').strip().strip('"'),
                    'key': canonical_drug_key(name),
                }
                all_rows.append(rec)
                by_key.setdefault(rec['key'], []).append(rec)

    return by_key, all_rows


def suggest_csv(section: str) -> str:
    return SECTION_TO_CSV.get(section.upper(), '')


def fuzzy_suggestions(
    excel_row: dict, all_csv: list[dict], top_n: int = 3
) -> list[tuple[float, dict]]:
    excel_tokens = norm_tokens(excel_row['match_name'])
    if not excel_tokens:
        return []

    scored: list[tuple[float, dict]] = []
    for rec in all_csv:
        csv_tokens = norm_tokens(rec['name'])
        if not csv_tokens:
            continue
        overlap = len(excel_tokens & csv_tokens)
        if overlap < 2:
            continue
        union = len(excel_tokens | csv_tokens)
        score = overlap / union if union else 0.0
        if score >= 0.35:
            scored.append((score, rec))

    scored.sort(key=lambda x: (-x[0], x[1]['name']))
    return scored[:top_n]


def format_candidates(candidates: list[dict]) -> str:
    parts = []
    for c in candidates:
        parts.append(f"{c['id']}|{c['file']}|{c['name']}")
    return ' ;; '.join(parts)


def format_fuzzy(suggestions: list[tuple[float, dict]]) -> str:
    parts = []
    for score, rec in suggestions:
        parts.append(f"{score:.2f}|{rec['id']}|{rec['file']}|{rec['name']}")
    return ' ;; '.join(parts)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)


def write_summary(
    path: Path,
    excel_products: list[dict],
    matched: list[dict],
    ambiguous: list[dict],
    missing: list[dict],
    csv_not_in_excel: list[dict],
) -> None:
    lines = [
        '# KEMSA Hospitals Tool → Odoo Product Match Summary',
        '',
        '## Totals',
        '',
        f'- Excel products: **{len(excel_products)}**',
        f'- Matched (single CSV hit): **{len(matched)}**',
        f'- Ambiguous (multiple CSV hits): **{len(ambiguous)}**',
        f'- Missing from product_variant: **{len(missing)}**',
        f'- CSV products not in Excel (informational): **{len(csv_not_in_excel)}**',
        '',
        'KEMSA codes are not stored in initializer CSVs today (`default_code` is empty).',
        'The matched report includes `recommended_default_code` for a future import.',
        '',
        '## Outcomes by Excel section',
        '',
        '| Section | Total | Matched | Ambiguous | Missing |',
        '|---------|------:|--------:|----------:|--------:|',
    ]

    section_stats: dict[str, dict[str, int]] = {}
    for p in excel_products:
        sec = p['section']
        section_stats.setdefault(sec, {'total': 0, 'matched': 0, 'ambiguous': 0, 'missing': 0})
        section_stats[sec]['total'] += 1

    for row in matched:
        section_stats[row['section']]['matched'] += 1
    for row in ambiguous:
        section_stats[row['section']]['ambiguous'] += 1
    for row in missing:
        section_stats[row['section']]['missing'] += 1

    for sec in sorted(section_stats):
        s = section_stats[sec]
        lines.append(
            f"| {sec} | {s['total']} | {s['matched']} | {s['ambiguous']} | {s['missing']} |"
        )

    lines.extend(
        [
            '',
            '## Top missing categories',
            '',
        ]
    )
    cat_counts = Counter(r['category'] for r in missing if r.get('category'))
    for cat, count in cat_counts.most_common(15):
        lines.append(f'- {count} — {cat}')

    lines.extend(['', '## CSV files with products not in Excel', ''])
    file_counts = Counter(r['csv_file'] for r in csv_not_in_excel)
    for fname, count in file_counts.most_common():
        lines.append(f'- {count} — `{fname}`')

    if ambiguous:
        lines.extend(['', '## Ambiguous rows (manual resolution needed)', ''])
        for row in ambiguous:
            lines.append(
                f"- `{row['kemsa_code']}` {row['excel_name'][:60]} "
                f"→ {row['candidate_count']} candidates"
            )

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text('\n'.join(lines) + '\n', encoding='utf-8')


def run(xlsx: Path, variant_dir: Path, out_dir: Path) -> int:
    if not xlsx.is_file():
        print(f'Excel not found: {xlsx}', file=sys.stderr)
        return 1
    if not variant_dir.is_dir():
        print(f'Variant dir not found: {variant_dir}', file=sys.stderr)
        return 1

    excel_products = parse_excel(xlsx)
    by_key, all_csv = load_csv_index(variant_dir)

    matched: list[dict] = []
    ambiguous: list[dict] = []
    missing: list[dict] = []
    missing_with_suggestions: list[dict] = []

    excel_keys: set[str] = set()

    for prod in excel_products:
        excel_keys.add(prod['key'])
        hits = by_key.get(prod['key'], [])

        if len(hits) == 1:
            hit = hits[0]
            matched.append(
                {
                    'kemsa_code': prod['kemsa_code'],
                    'recommended_default_code': prod['kemsa_code'],
                    'excel_name': prod['name'],
                    'match_name': prod['match_name'],
                    'is_tracer': prod['is_tracer'],
                    'section': prod['section'],
                    'category': prod['category'],
                    'pack_size': prod['pack_size'],
                    'excel_price': prod['price'],
                    'excel_row': prod['excel_row'],
                    'odoo_id': hit['id'],
                    'odoo_name': hit['name'],
                    'csv_file': hit['file'],
                    'odoo_lst_price': hit['lst_price'],
                    'odoo_default_code': hit['default_code'],
                }
            )
        elif len(hits) > 1:
            ambiguous.append(
                {
                    'kemsa_code': prod['kemsa_code'],
                    'recommended_default_code': prod['kemsa_code'],
                    'excel_name': prod['name'],
                    'match_name': prod['match_name'],
                    'is_tracer': prod['is_tracer'],
                    'section': prod['section'],
                    'category': prod['category'],
                    'excel_row': prod['excel_row'],
                    'candidate_count': len(hits),
                    'candidates': format_candidates(hits),
                }
            )
        else:
            suggestions = fuzzy_suggestions(prod, all_csv)
            suggested_csv = suggest_csv(prod['section'])
            row = {
                'kemsa_code': prod['kemsa_code'],
                'recommended_default_code': prod['kemsa_code'],
                'excel_name': prod['name'],
                'match_name': prod['match_name'],
                'is_tracer': prod['is_tracer'],
                'section': prod['section'],
                'category': prod['category'],
                'pack_size': prod['pack_size'],
                'excel_price': prod['price'],
                'excel_row': prod['excel_row'],
                'suggested_csv_file': suggested_csv,
                'fuzzy_suggestions': format_fuzzy(suggestions),
            }
            missing.append(row)
            missing_with_suggestions.append(row)

    csv_not_in_excel: list[dict] = []
    for rec in all_csv:
        if rec['key'] and rec['key'] not in excel_keys:
            csv_not_in_excel.append(
                {
                    'odoo_id': rec['id'],
                    'odoo_name': rec['name'],
                    'csv_file': rec['file'],
                    'odoo_lst_price': rec['lst_price'],
                    'odoo_default_code': rec['default_code'],
                }
            )

    out_dir.mkdir(parents=True, exist_ok=True)

    write_csv(
        out_dir / 'matched.csv',
        [
            'kemsa_code',
            'recommended_default_code',
            'excel_name',
            'match_name',
            'is_tracer',
            'section',
            'category',
            'pack_size',
            'excel_price',
            'excel_row',
            'odoo_id',
            'odoo_name',
            'csv_file',
            'odoo_lst_price',
            'odoo_default_code',
        ],
        matched,
    )
    write_csv(
        out_dir / 'ambiguous.csv',
        [
            'kemsa_code',
            'recommended_default_code',
            'excel_name',
            'match_name',
            'is_tracer',
            'section',
            'category',
            'excel_row',
            'candidate_count',
            'candidates',
        ],
        ambiguous,
    )
    write_csv(
        out_dir / 'missing.csv',
        [
            'kemsa_code',
            'recommended_default_code',
            'excel_name',
            'match_name',
            'is_tracer',
            'section',
            'category',
            'pack_size',
            'excel_price',
            'excel_row',
            'suggested_csv_file',
        ],
        missing,
    )
    write_csv(
        out_dir / 'missing_with_suggestions.csv',
        [
            'kemsa_code',
            'recommended_default_code',
            'excel_name',
            'match_name',
            'is_tracer',
            'section',
            'category',
            'pack_size',
            'excel_price',
            'excel_row',
            'suggested_csv_file',
            'fuzzy_suggestions',
        ],
        missing_with_suggestions,
    )
    write_csv(
        out_dir / 'csv_not_in_excel.csv',
        ['odoo_id', 'odoo_name', 'csv_file', 'odoo_lst_price', 'odoo_default_code'],
        csv_not_in_excel,
    )
    write_summary(
        out_dir / 'summary.md',
        excel_products,
        matched,
        ambiguous,
        missing,
        csv_not_in_excel,
    )

    print(f'Excel products: {len(excel_products)}')
    print(f'Matched: {len(matched)}')
    print(f'Ambiguous: {len(ambiguous)}')
    print(f'Missing: {len(missing)}')
    print(f'CSV not in Excel: {len(csv_not_in_excel)}')
    print(f'Reports written to {out_dir}')
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description='Match KEMSA hospitals tool products to Odoo initializer CSVs.'
    )
    parser.add_argument('--xlsx', type=Path, default=DEFAULT_XLSX)
    parser.add_argument('--variant-dir', type=Path, default=DEFAULT_VARIANT_DIR)
    parser.add_argument('--out', type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()
    return run(args.xlsx, args.variant_dir, args.out)


if __name__ == '__main__':
    raise SystemExit(main())
